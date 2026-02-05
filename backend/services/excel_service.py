"""
Service for updating Excel files with comparison results.
Safely handles Excel file operations and preserves existing formatting.
"""
import os
import sys
from datetime import datetime
from typing import List, Optional, Tuple, Dict
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage


sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


from models.schemas import FileDiff
from utils.file_utils import path_exists

def sheet_name_from_component(component_name: str,
                              default: str = "Diff Screenshots") -> str:
    """
    Build a valid Excel sheet name from a component name.
    - Removes invalid characters: [] : * ? / \
    - Trims to 31 chars (Excel limit)
    - Falls back to default if empty/invalid
    """
    name = (component_name or "").strip()
    if not name:
        return default

    invalid = set('[]:*?/\\')
    cleaned = ''.join(c for c in name if c not in invalid).strip()
    if not cleaned:
        return default

    # Excel sheet name max length = 31
    if len(cleaned) > 31:
        cleaned = cleaned[:31]

    return cleaned

def check_excel_open(excel_path: str) -> bool:
  """
  Check if Excel file is open by attempting to open it in write mode.

  Args:
      excel_path: Path to Excel file

  Returns:
      True if file appears to be open, False otherwise
  """
  if not path_exists(excel_path):
    return False

  try:
    # Try to open file in append mode - will fail if file is open
    with open(excel_path, "r+b"):
      return False
  except (PermissionError, IOError):
    return True



def add_diff_image_to_excel(
    excel_path: str,
    file_name: str,
    image_file_path: str,
    component_name: str = "",
    sheet_name: str = None,
    server_name: str = "",
) -> Tuple[bool, str, int]:
    """
    Add a screenshot image to the given Excel file in a separate sheet.

    - One sheet per component (based on component_name)
    - Every call appends a new block (meta row + image row) at the *bottom*
      of the sheet, so repeated "Capture all" runs never overlap old images.
    """
        # Check if Excel is open
    if check_excel_open(excel_path):
        return False, "Please close Excel file first", 0

    try:
        # Load existing workbook or create new one
        if path_exists(excel_path):
            workbook = load_workbook(excel_path)
        else:
            workbook = Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

        # Decide which sheet to use: one sheet per component
        if sheet_name is None:
            sheet_name = sheet_name_from_component(component_name)

        # Get or create the target sheet
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # --- Ensure server header row at top (once) ---

        if server_name:
            first_val = str(sheet["A1"].value or "")
            if not first_val.startswith("Server:"):
                # If row 1 already has data, push everything down by one row
                if sheet.max_row > 0 and any(c.value is not None for c in sheet[1]):
                    sheet.insert_rows(1)
                sheet["A1"] = f"Server: {server_name}"

        # --- Ensure column headers below server row ---

        header_row_index = 2 if str(sheet["A1"].value or "").startswith("Server:") else 1
        first_header_cell = sheet.cell(row=header_row_index, column=1).value

        if first_header_cell != "Component Name":
            headers = ["Component Name", "File Name", "Timestamp", "Screenshot"]
            for col, value in enumerate(headers, start=1):
                sheet.cell(row=header_row_index, column=col, value=value)

        # --- Decide where to put the next screenshot ---

        # Always append *after* the last used row in the sheet.
        last_row = sheet.max_row or 1

        # Leave a small gap under the last used row
        GAP_ROWS = 2
        meta_row = last_row + GAP_ROWS     # row for component/file/timestamp
        image_row = meta_row + 1           # row where the image is anchored

        # Meta info
        sheet.cell(row=meta_row, column=1, value=component_name)
        sheet.cell(row=meta_row, column=2, value=file_name)
        sheet.cell(
            row=meta_row,
            column=3,
            value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

        # --- Load and scale image (WIDTH only) ---
        img = XLImage(image_file_path)

        orig_width = img.width or 1
        orig_height = img.height or 1

        MAX_WIDTH_PX = 1800
        if orig_width > MAX_WIDTH_PX:
            scale = MAX_WIDTH_PX / float(orig_width)
        else:
            scale = 1.0

        new_width = int(orig_width * scale)
        new_height = int(orig_height * scale)

        img.width = new_width
        img.height = new_height

        # Anchor image at column D in the chosen row
        img.anchor = f"D{image_row}"
        sheet.add_image(img)

        # --- Layout tuning ---

        # Make column D roughly match the image width (Excel units ≈ px / 7.5)
        desired_width = new_width / 7.5
        col_dim = sheet.column_dimensions["D"]
        current_width = col_dim.width or 0
        if desired_width > current_width:
            col_dim.width = desired_width

        # Set the height of the image row so the image fits within that row.
        try:
            sheet.row_dimensions[image_row].height = new_height * 0.75
        except Exception:
            pass

        # Ensure this image row counts for max_row next time
        if sheet.cell(row=image_row, column=4).value is None:
            sheet.cell(row=image_row, column=4, value="")

        workbook.save(excel_path)
        return True, "Diff image added clearly", 1

    except PermissionError:
        return False, "Please close Excel file first", 0
    except Exception as exc:
        return False, f"Error writing screenshot to Excel: {exc}", 0
    
def update_excel_file(
    excel_path: str,
    file_diffs: List[FileDiff],
    comments: Optional[Dict[str, Dict[str, str]]] = None,
    sheet_name: str = "Configuration Comparison"
) -> Tuple[bool, str, int]:
  """
  Update Excel file with comparison results.
  Preserves existing formatting and adds new rows.

  Args:
      excel_path: Path to Excel file
      file_diffs: List of FileDiff objects
      sheet_name: Name of the sheet to update

  Returns:
      Tuple of (success, message, updated_rows)
  """
  # Check if Excel is open
  if check_excel_open(excel_path):
    return False, "Please close Excel file first", 0

  try:
    # Load existing workbook or create new one
    if path_exists(excel_path):
      try:
        workbook = load_workbook(excel_path)
      except Exception as e:
        return False, f"Error opening Excel file: {str(e)}", 0
    else:
      workbook = Workbook()
      # Remove default sheet if it exists
      if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    # Get or create the target sheet
    if sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
    else:
      sheet = workbook.create_sheet(sheet_name)
      # Add headers if new sheet
      headers = ["Component Name", "Config File Name", "Changes", "Date of Comparison"]
      sheet.append(headers)
      # Style headers
      header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
      header_font = Font(bold=True, color="FFFFFF")
      for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Find the last row with data
    max_row = sheet.max_row
    start_row = max_row + 1

    # Determine if we will write per-key semantic rows (requires expanded headers)
    will_expand = any(
      (getattr(fd, 'semantic_diff', {}) or {}).get('changes')
      for fd in file_diffs if getattr(fd, 'has_changes', False)
    )
    if will_expand:
      try:
        first_row_values = [cell.value for cell in sheet[1]]
      except Exception:
        first_row_values = []
      if 'Key/Change' not in first_row_values:
        new_headers = [
          "Component Name",
          "Config File Name",
          "Key/Change",
          "Old Value",
          "New Value",
          "Comment",
          "Date of Comparison",
        ]
        for i, h in enumerate(new_headers, start=1):
          sheet.cell(row=1, column=i).value = h

    updated_rows = 0
    comparison_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for file_diff in file_diffs:
      if not file_diff.has_changes:
        continue

      added = sum(1 for line in file_diff.diff_lines if line.line_type == "added")
      removed = sum(1 for line in file_diff.diff_lines if line.line_type == "removed")

      change_summary = "No changes detected"
      if added > 0 or removed > 0:
        parts = []
        if added > 0:
          parts.append(f"{added} line(s) added")
        if removed > 0:
          parts.append(f"{removed} line(s) removed")
        change_summary = "; ".join(parts)

      semantic = getattr(file_diff, 'semantic_diff', None) or {}
      changes = semantic.get('changes', []) if isinstance(semantic, dict) else []

      if changes:
        for ch in changes:
          key = ch.get('key') or (ch.get('old_key') + ' -> ' + ch.get('new_key', ''))
          comment = None
          if comments and isinstance(comments, dict):
            file_comments = comments.get(getattr(file_diff, 'new_path', '') or '') or {}
            comment = file_comments.get(ch.get('key')) or file_comments.get(ch.get('new_key'))

          row_data = [
            file_diff.component_name,
            file_diff.file_name,
            key,
            str(ch.get('old_value')),
            str(ch.get('new_value')),
            comment or '',
            comparison_date,
          ]
          sheet.append(row_data)
          updated_rows += 1
      else:
        row_data = [
          file_diff.component_name,
          file_diff.file_name,
          change_summary,
          comparison_date,
        ]
        sheet.append(row_data)
        updated_rows += 1

    message = f"Excel updated successfully. Added {updated_rows} row(s)."
    workbook.save(excel_path)
    return True, message, updated_rows

  except PermissionError:
    return False, "Please close Excel file first", 0
  except Exception as e:
    return False, f"Error updating Excel: {str(e)}", 0


def write_changes_to_excel(
    excel_path: str,
    changes: List[Dict[str, str]],
    sheet_name: str = "Reviewed Changes"
) -> Tuple[bool, str, int]:
  """
  Write reviewed changes to Excel file.
  Creates or appends to sheet with columns: Component Name, File Name, Changed Line, Comment

  Args:
      excel_path: Path to Excel file
      changes: List of change dicts with keys: componentName, fileName, changedLine, comment
      sheet_name: Name of the sheet to write to

  Returns:
      Tuple of (success, message, written_rows)
  """
  if check_excel_open(excel_path):
    return False, "Please close Excel file first", 0

  try:
    if path_exists(excel_path):
      try:
        workbook = load_workbook(excel_path)
      except Exception as e:
        return False, f"Error opening Excel file: {str(e)}", 0
    else:
      workbook = Workbook()
      if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    if sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
    else:
      sheet = workbook.create_sheet(sheet_name)
      headers = ["Component Name", "File Name", "Changed Line", "Comment"]
      sheet.append(headers)
      header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
      header_font = Font(bold=True, color="FFFFFF")
      for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    written_rows = 0
    for change in changes:
      row_data = [
        change.get('componentName', ''),
        change.get('fileName', ''),
        change.get('changedLine', ''),
        change.get('comment', ''),
      ]
      sheet.append(row_data)
      written_rows += 1

    # Auto-adjust column widths
    for col_num, column in enumerate(sheet.columns, 1):
      max_length = 0
      column_letter = get_column_letter(col_num)
      for cell in column:
        try:
          if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
        except Exception:
          pass
      adjusted_width = min(max_length + 2, 50)  # Cap at 50
      sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(excel_path)
    message = f"Changes written to Excel successfully. Added {written_rows} row(s)."
    return True, message, written_rows

  except PermissionError:
    return False, "Please close Excel file first", 0
  except Exception as e:
    return False, f"Error writing to Excel: {str(e)}", 0
  
def write_delta_groups_to_excel(
    excel_path: str,
    database_name: str,
    groups: List[Dict],
    sheet_name: str = None,
) -> Tuple[bool, str, int]:
    """
    Write delta groups into Excel in a matrix layout:

        B2: <DatabaseName>

        B4: TableScripts   C4: StoredProcedures   D4: Functions   E4: Index   ...

        Under each folder name column, list the .sql files in that folder.

    One sheet per database. If the sheet already exists, it is cleared
    and rewritten with the new content.
    """
    if check_excel_open(excel_path):
        return False, "Please close Excel file first", 0

    try:
        # Load or create workbook
        if path_exists(excel_path):
            try:
                workbook = load_workbook(excel_path)
            except Exception as e:
                return False, f"Error opening Excel file: {str(e)}", 0
        else:
            workbook = Workbook()
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

        # Decide sheet name: default to database_name (sanitized to Excel rules)
        if not sheet_name:
            # Reuse the sheet_name_from_component helper, just with db name
            sheet_name = sheet_name_from_component(database_name or "Database")

        # If sheet already exists, remove it (we want a fresh layout each time)
        if sheet_name in workbook.sheetnames:
            ws_old = workbook[sheet_name]
            workbook.remove(ws_old)

        sheet = workbook.create_sheet(sheet_name)

        # Layout constants
        DB_ROW = 2          # row where database name goes
        HEADER_ROW = 4      # row where folder names go
        START_COL = 2       # column B (1-based index)

        # 1) Database name at B2
        sheet.cell(row=DB_ROW, column=START_COL, value=database_name)

        # 2) Folder names in row 4, starting at B4
        #    We use the order given by 'groups'
        written_rows = 0
        for idx, group in enumerate(groups or []):
            col = START_COL + idx
            header_cell = sheet.cell(row=HEADER_ROW, column=col, value=group.get("name", ""))

            # Simple header styling
            header_cell.font = Font(bold=True, color="000000")
            header_cell.alignment = Alignment(horizontal="center", vertical="center")

            # 3) Files listed under each folder name
            row = HEADER_ROW + 1
            for f in group.get("files", []):
                sheet.cell(row=row, column=col, value=f.get("file_name", ""))
                row += 1
                written_rows += 1

        # Auto-fit column widths based on longest value per column
        for col_num, column in enumerate(sheet.columns, 1):
            max_length = 0
            col_letter = get_column_letter(col_num)
            for cell in column:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            if max_length > 0:
                sheet.column_dimensions[col_letter].width = min(max_length + 2, 60)

        workbook.save(excel_path)
        msg = (
            f"Delta groups written for database '{database_name}' "
            f"to sheet '{sheet_name}'. Added {written_rows} file name(s)."
        )
        return True, msg, written_rows

    except PermissionError:
        return False, "Please close Excel file first", 0
    except Exception as e:
        return False, f"Error writing delta groups to Excel: {str(e)}", 0